import shutil
import os
import re
import time
import win32com.client as win32
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import pandas as pd

today = datetime.now().strftime("%Y-%m-%d")
month = datetime.now().strftime("%m")
day = datetime.now().strftime("%d")


# 刷新表格数据包括powerquery里面生成的表
def excel_refresh(file_path):
    # file_path = r"D:\移动终端广分互联网\小时达\管理单元及门店管理\管理单元&门店明细.xlsx"

    # 打开 Excel 应用程序
    excel_app = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        excel_app.Visible = False  # 隐藏 Excel 窗口

        # 打开工作簿
        workbook = excel_app.Workbooks.Open(file_path)

        # 刷新所有数据连接和 Power Query 查询
        workbook.RefreshAll()

        # 等待刷新完成
        excel_app.CalculateUntilAsyncQueriesDone()

        # 保存并关闭工作簿
        workbook.Save()
        workbook.Close()

    finally:
        # 退出 Excel 应用程序
        excel_app.Quit()


# 分别更新管理单元&门店明细表
def excel_update_data():
    # 指定 Excel 文件的路径
    source_file_path = fr"D:\移动终端广分互联网\小时达\管理单元及门店管理\管理单元&门店明细.xlsx"  # 使用原始字符串（r）来避免转义问题
    gldy_file_path = fr"D:\移动终端广分互联网\小时达\管理单元及门店管理\{today}\最新-管理单元&超管号未通过邀请明细{month}{day}.xlsx"
    md_file_path = fr"D:\移动终端广分互联网\小时达\管理单元及门店管理\{today}\最新-门店&管理号未通过邀请明细{month}{day}.xlsx"

    source_wb = openpyxl.load_workbook(source_file_path)
    gldy_wb = openpyxl.load_workbook(gldy_file_path)
    md_wb = openpyxl.load_workbook(md_file_path)

    # 获取指定工作表
    source_gldy_ws = source_wb["管理单元信息"]
    source_md_ws = source_wb["门店信息"]
    target_gldy_ws = gldy_wb.worksheets[0]
    target_md_ws = md_wb.worksheets[0]

    # 复制源工作表内容到目标工作表
    for row in source_gldy_ws.iter_rows(min_row=2):     # 管理单元表格更新
        for cell in row:
            target_gldy_ws[cell.coordinate].value = cell.value

    gldy_wb.save(gldy_file_path)

    for row in source_md_ws.iter_rows(min_row=2):  # 门店表格更新
        for cell in row:
            target_md_ws[cell.coordinate].value = cell.value

    md_wb.save(md_file_path)


# 将小时达入驻情况表内的公式转值并附上对应数量至表格标题
def excel_conversion():
    # 指定 Excel 文件的路径
    file_path = fr"D:\移动终端广分互联网\小时达\管理单元及门店管理\{today}\小时达入驻情况.xlsx"  # 使用原始字符串（r）来避免转义问题
    file_path_backup = fr"D:\移动终端广分互联网\小时达\管理单元及门店管理\{today}\小时达入驻情况{month}{day}.xlsx"  # 使用原始字符串（r）来避免转义问题

    # 打开 Excel 文件，设置 data_only=True 以便读取公式计算后的值
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(file_path)
    except FileNotFoundError:
        wb = openpyxl.load_workbook(file_path_backup, data_only=True)
        print(file_path_backup)

    # 遍历每个工作表
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # 遍历每个单元格
        for row in ws.iter_rows():
            for cell in row:    # 将单元格的值替换为公式计算后的值（如果有公式）
                if cell.data_type == 'f':  # 检查单元格是否包含公式
                    cell.value = cell.value  # 将公式替换为计算后的数值

    # 修改每个工作表名称，将门店数附在名称后面
    sheet_name = wb.sheetnames
    for sn in sheet_name:
        ws = wb[sn]
        sn = re.sub(r"\d*$", "", sn)    # 清除后面带有的数字
        # 判断汇总表定位合计，找到合计数
        if sn == "省区管理单元" or "省区门店":
            search_value = "合计"
            # 仅遍历A列（即第一列）
            for cell in ws.iter_rows(min_col=1, max_col=1):
                if cell[0].value == search_value:  # cell[0] 是 A 列的单元格
                    sum_cell = ws.cell(cell[0].row, cell[0].column + 2)    # 合计单元格向右偏移2列即为合计数
                    ws.title = sn + str(sum_cell.value)
                    print(ws.title)
                    break
        # 判断是否地市表，若是在表面加上行数即门店数
        if len(sn) == 2:    # 一般地市都是2个字
            ws.title = sn + str(ws.max_row - 1)
            print(ws.title)

    # 保存修改后的文件
    wb.save(file_path_backup)  # 保存为新文件或覆盖原文件


# 将未能匹配的管理单元&门店复制到匹配表待补充
def city_need_add():
    # 加载明细表和需要补充的地市表
    dt_path = r"D:\移动终端广分互联网\小时达\管理单元及门店管理\管理单元&门店明细.xlsx"
    sup_path = r"D:\移动终端广分互联网\小时达\管理单元及门店管理\数据源\管理单元&门店地市匹配.xlsx"
    dt_wb = openpyxl.load_workbook(dt_path)
    sup_wb = openpyxl.load_workbook(sup_path)

    # 分管理单元和门店分别进行补充
    dt_sheet_name = ["管理单元信息", "门店信息"]
    sup_sheet_name = ["管理单元地市匹配", "门店地市匹配"]

    for dt, sup in zip(dt_sheet_name, sup_sheet_name):
        # 获取明细工作表和匹配工作表
        source_ws = dt_wb[dt]
        target_ws = sup_wb[sup]

        # 获取匹配表的最大行，准备在末尾插入数据
        target_max_row = target_ws.max_row

        # 遍历明细表的所有行，筛选第三列为空白的行
        for row in source_ws.iter_rows(min_row=2):  # 从第二行开始，假设第一行是表头
            third_col_value = row[2].value  # 第三列数据
            if third_col_value is None or third_col_value == "":  # 判断是否为空
                # 获取前两列的数据
                first_col_value = row[0].value
                second_col_value = row[1].value

                # 在目标表最后一行后面插入前两列数据
                target_ws.cell(row=target_max_row + 1, column=1, value=first_col_value)
                target_ws.cell(row=target_max_row + 1, column=2, value=second_col_value)

                # 更新目标表的最大行
                target_max_row += 1

    # 保存文件
    sup_wb.save(r"D:\移动终端广分互联网\小时达\管理单元及门店管理\数据源\管理单元&门店地市匹配.xlsx")  # 保存为新文件，或覆盖原文件


# 复制一个当天的文件夹，并将改日期名称
def copy_folder():
    # 定义源文件夹路径和目标文件夹路径
    source_folder = r"D:\移动终端广分互联网\小时达\管理单元及门店管理\模板"  # 替换为你的源文件夹路径
    source_file = r"D:\移动终端广分互联网\小时达\管理单元及门店管理\小时达入驻情况.xlsx"
    target_folder = os.path.join(os.path.dirname(source_folder), today)  # 在同目录下创建目标文件夹

    # 复制模板文件夹并重命名为当天日期
    shutil.copytree(source_folder, target_folder)

    # 复制小时达入驻情况.xlsx到新文件夹
    shutil.copy(source_file, target_folder)

    # 更新表格名称
    file_keywords = ["最新-管理单元&超管号未通过邀请明细", "最新-门店&管理号未通过邀请明细"]    # 要匹配的关键词
    # 正则表达式，匹配以数字结尾的文件名
    date_pattern = re.compile(r"(\d{4})")  # 匹配文件名4位数字月日

    # 遍历文件夹中的文件
    for file_name in os.listdir(target_folder):
        # 构建完整文件路径
        old_file_path = os.path.join(target_folder, file_name)

        # 仅处理文件，排除文件夹
        if os.path.isfile(old_file_path):
            # 检查文件名是否包含指定关键词
            if any(keyword in file_name for keyword in file_keywords):
                # 使用正则表达式匹配文件名中的日期部分
                today_md = datetime.now().strftime("%m%d")
                new_file_name = date_pattern.sub(today_md, file_name)
                new_file_path = os.path.join(target_folder, new_file_name)

                # 重命名文件
                os.rename(old_file_path, new_file_path)
                print(f"文件已重命名为: {new_file_name}")


# 匹配B2B合同签署状态
def deal_excel(substring, file_path):
    # 定位到需要处理的文件
    deal_date = today  # 需要编辑的日期
    today_folder = rf'D:\移动终端广分互联网\小时达\管理单元及门店管理\{deal_date}'  # 所处理文件路径

    # 获取文件夹下的文件名
    file_names = os.listdir(today_folder)
    print(f"文件夹下的文件列表: {file_names}")

    # 整理出所需处理文件的文件路径
    file_name = [file for file in file_names if substring in file]
    print(f"匹配到的文件名: {file_name}")

    if not file_name:
        print(f"未找到包含 '{substring}' 的文件")
        return
    elif len(file_name) > 1:
        file_name.pop(0)

    excel_path = os.path.join(today_folder, file_name[0])
    print(f"尝试访问的文件路径: {excel_path}")

    # 检查文件是否存在
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return

    # 读取 Excel 文件
    df_B2B = pd.read_excel(file_path, sheet_name='合同推送情况')

    if substring == '最新-管理单元&超管号未通过邀请明细':
        hand_sheet = '管理单元通过&超管号明细'
        df_hand = pd.read_excel(excel_path, sheet_name=hand_sheet)

        # 将 df_B2B 的 '管理单元id' 列转换为字符串类型，以便与 df_hand 的 '管理单元id' 匹配
        df_hand['管理单元id'] = df_hand['管理单元id'].astype(str)
        df_B2B['管理单元id'] = df_B2B['管理单元id'].astype(str)

        # 使用 merge 进行匹配
        merged_df = pd.merge(
            df_hand,
            df_B2B[['管理单元id', '判定状态']],
            how='left',
            left_on='管理单元id',
            right_on='管理单元id'
        )

    elif substring == '最新-门店&管理号未通过邀请明细':
        hand_sheet = '门店通过&管理号明细'
        df_hand = pd.read_excel(excel_path, sheet_name=hand_sheet)

        # 使用 merge 进行匹配
        merged_df = pd.merge(
            df_hand,
            df_B2B[['管理单元名称', '判定状态']].drop_duplicates(subset='管理单元名称', keep='first'),  # 去重，只保留第一次匹配
            how='left',
            left_on='上级管理单元',
            right_on='管理单元名称'
        )

    # 小时达入驻情况表有点不一样，所以要单独处理
    elif substring == '小时达入驻情况':
        # 读取 Excel 文件的所有 sheet 名称
        with pd.ExcelFile(excel_path) as excel:
            sheet_names = excel.sheet_names  # 获取所有 sheet 名称

        # 定义正则表达式模式
        pattern = re.compile(r'省区管理单元\d')
        # 筛选出符合条件的 sheet 名称
        matched_sheet = [name for name in sheet_names if pattern.match(name)][0]

        df_hand = pd.read_excel(excel_path, sheet_name=matched_sheet)

        # 处理地市名称匹配
        df_hand['匹配地市'] = df_hand['地市'].iloc[0:21] + '市'  # 在目标文件中添加“市”字以便匹配

        # 统计每个地市的“待审批”数量
        signed_counts = []
        for city in df_hand['匹配地市'].iloc[0:21]:  # 只对前21行进行操作
            count = df_B2B[(df_B2B['营业执照-所属地市'] == city) & (df_B2B['判定状态'] == '待审批')].shape[0]
            signed_counts.append(count)

        # 22行和30行放总计，其余中间行空着
        total_counts = sum(signed_counts)
        signed_counts = signed_counts + [total_counts] + [None] * (len(df_hand) - 23) + [total_counts]

        # 将统计结果写入目标文件的 H 列
        df_hand['已签署'] = pd.Series(signed_counts)

        # 使用 openpyxl 加载原始文件以保留格式
        book = openpyxl.load_workbook(excel_path)
        sheet = book[matched_sheet]

        # 在 H 列插入新列
        sheet.insert_cols(8)  # 在第8列（H列）插入新列
        sheet.cell(row=1, column=8, value='已签署')  # 设置标题

        # 写入统计结果
        for i, count in enumerate(signed_counts, start=2):  # 从第2行开始写入
            sheet.cell(row=i, column=8, value=count)

        # 设置 H 列格式与 G 列相同
        for row in sheet.iter_rows(min_col=7, max_col=8, min_row=1, max_row=sheet.max_row):
            g_cell = row[0]  # G 列单元格
            h_cell = row[1]  # H 列单元格
            copy_style(g_cell, h_cell)  # 复制样式

        G_width = sheet.column_dimensions['G'].width  # 获取G列宽
        sheet.column_dimensions['H'].width = G_width    # 将H列宽设为与G列相同

        # 保存文件
        book.save(excel_path)
        print(f"处理完成，结果已保存到原始文件: {excel_path}")
        return

    # 处理匹配结果，将 NaN 替换为 '未推送'
    merged_df['判定状态'] = merged_df['判定状态'].fillna('未推送')

    # 在 df_hand 的 C/D 列后插入新列
    df_hand.insert(3, 'B2B签署情况', merged_df['判定状态'])

    # 使用 openpyxl 加载原始文件以保留格式
    book = openpyxl.load_workbook(excel_path)

    # 获取目标工作表
    sheet_name = hand_sheet
    if sheet_name not in book.sheetnames:
        print(f"未找到工作表: {sheet_name}")
        return

    sheet = book[sheet_name]

    # 清空内容（保留标题行）
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)  # 从第二行开始删除

    # 将新标题和数据写入目标工作表
    for i, row in enumerate(dataframe_to_rows(df_hand, index=False, header=True)):  # 包括标题
        for j, value in enumerate(row):
            sheet.cell(row=i + 1, column=j + 1, value=value)  # 从标题开始写入

    # 设置标题行样式
    header_row = sheet[1]  # 第一行是标题行
    for cell in header_row:
        set_cell_format(sheet, cell.coordinate, fill_color="FFFF00", border=True)   # 底色黄色，全边框
    set_cell_format(sheet, "D1", fill_color="F4B084", border=True)  # 底色橙色，全边框

    # 保存文件
    book.save(excel_path)
    print(f"处理完成，结果已保存到原始文件: {excel_path}")


# 改变单元格样式
def set_cell_format(sheet, cell_position, fill_color=None, font_color='000000', border=False):
    """
    设置单元格的格式

    :param sheet: Excel工作表对象
    :param cell_position: 要修改的单元格位置（例如 'A1'）
    :param fill_color: 底色（例如 'FFFF00'），如果不设置，传入 None
    :param font_color: 字体颜色（例如 'FF0000'），如果不设置，传入 None
    :param border: 是否设置全框线（True/False）
    """
    # 获取目标单元格
    cell = sheet[cell_position]

    # 设置底色
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # 设置字体颜色
    if font_color:
        cell.font = Font(color=font_color)

    # 设置全框线
    if border:
        border_style = Side(border_style="thin", color="000000")  # 细黑色边框
        cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)


# 复制样式函数
def copy_style(source_cell, target_cell):
    """
    复制源单元格的样式到目标单元格，包括填充、字体、边框、对齐方式和列宽。
    """
    # 复制填充样式
    if source_cell.fill:
        target_cell.fill = PatternFill(
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color,
            fill_type=source_cell.fill.fill_type
        )

    # 复制字体样式
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )

    # 复制边框样式
    if source_cell.border:
        target_cell.border = Border(
            left=Side(border_style=source_cell.border.left.border_style, color=source_cell.border.left.color),
            right=Side(border_style=source_cell.border.right.border_style, color=source_cell.border.right.color),
            top=Side(border_style=source_cell.border.top.border_style, color=source_cell.border.top.color),
            bottom=Side(border_style=source_cell.border.bottom.border_style, color=source_cell.border.bottom.color)
        )

    # 复制对齐方式
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,  # 水平对齐（如 'center'）
            vertical=source_cell.alignment.vertical,  # 垂直对齐（如 'center'）
            wrap_text=source_cell.alignment.wrap_text,  # 是否自动换行
            shrink_to_fit=source_cell.alignment.shrink_to_fit,  # 是否缩小字体以适应
            indent=source_cell.alignment.indent,  # 缩进
            text_rotation=source_cell.alignment.text_rotation  # 文本旋转角度
        )

